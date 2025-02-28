import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
import uuid
import zipfile
import base64
from urllib.parse import urljoin, urlparse
from PIL import Image
from io import BytesIO
import time
import openpyxl
from openpyxl.drawing.image import Image as XLImage

def get_site_logo(url):
    """
    Scrape a website to find and download its logo image.
    Returns the logo info or None if no logo was found.
    """
    # Make sure URL has proper format
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    
    try:
        # Send a request to get the website content
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()  # Raise exception for HTTP errors
        
        # Parse the HTML content
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Look for common logo patterns
        logo_candidates = []
        
        # 1. Check for link tags with 'icon' or 'logo' in the rel attribute
        for link in soup.find_all('link'):
            rel = link.get('rel', [])
            if isinstance(rel, list):
                rel = ' '.join(rel).lower()
            else:
                rel = str(rel).lower()
                
            if ('icon' in rel or 'logo' in rel) and link.get('href'):
                logo_candidates.append((link.get('href'), 1))  # Lower priority
        
        # 2. Look for images with 'logo' in the class, id, or alt attribute
        for img in soup.find_all('img'):
            score = 0
            for attr in ['class', 'id', 'alt', 'src']:
                value = img.get(attr, '')
                if isinstance(value, list):
                    value = ' '.join(value)
                else:
                    value = str(value)
                    
                if 'logo' in value.lower():
                    score += 2
                
            if score > 0 and img.get('src'):
                logo_candidates.append((img.get('src'), score + 2))  # Higher priority
        
        # If no logo candidates found, look for the first image in the header
        if not logo_candidates:
            header = soup.find('header')
            if header:
                img = header.find('img')
                if img and img.get('src'):
                    logo_candidates.append((img.get('src'), 3))  # Medium priority
        
        # Sort by priority (highest first)
        logo_candidates.sort(key=lambda x: x[1], reverse=True)
        
        # Process logo candidates
        for img_url, _ in logo_candidates:
            # Convert relative URLs to absolute URLs
            img_url = urljoin(url, img_url)
            
            try:
                # Download the image
                img_response = requests.get(img_url, headers=headers, timeout=10)
                img_response.raise_for_status()
                
                # Check if it's a valid image
                img = Image.open(BytesIO(img_response.content))
                
                # Convert to RGB if needed (for PNG with transparency)
                if img.mode == 'RGBA':
                    # Create a white background
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    # Paste using alpha channel as mask
                    background.paste(img, mask=img.split()[3])
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Generate a unique filename
                domain = urlparse(url).netloc
                # Always save as JPG for consistency
                filename = f"{domain.replace('.', '_')}_{uuid.uuid4().hex[:8]}.jpg"
                
                # Create images directory if it doesn't exist
                os.makedirs('logos', exist_ok=True)
                
                # Save the image as JPG
                img_path = os.path.join('logos', filename)
                img.save(img_path, 'JPEG', quality=85)
                
                # Save the image data to memory as well
                img_byte_arr = BytesIO()
                img.save(img_byte_arr, format='JPEG', quality=85)
                img_data = img_byte_arr.getvalue()
                
                # Return all the image information
                return {
                    'path': img_path,
                    'filename': filename,
                    'data': img_data,
                    'format': 'jpg',
                    'pil_image': img
                }
            
            except Exception as e:
                continue  # Try the next candidate if this one fails
        
        return None  # No valid logo found
        
    except Exception as e:
        st.error(f"Error processing {url}: {str(e)}")
        return None

def create_excel_with_embedded_images(df, logo_info_dict, url_column):
    """
    Create an Excel file with the logos embedded in cells.
    """
    # Create a new Excel workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logo Extraction Results"
    
    # Add headers
    headers = list(df.columns) + ["Logo Image"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    # Add data and images
    for row_num, (_, row) in enumerate(df.iterrows(), 2):
        # Add regular cell data
        for col_num, col_name in enumerate(df.columns, 1):
            ws.cell(row=row_num, column=col_num).value = row[col_name]
        
        # Add image if available
        url = row[url_column]
        if url in logo_info_dict and logo_info_dict[url]:
            logo_path = logo_info_dict[url]['path']
            try:
                # Place image in the last column
                img = XLImage(logo_path)
                # Resize the image to fit in a cell
                img.width = 150
                img.height = 75
                ws.add_image(img, f"{chr(65 + len(df.columns))}{row_num}")
                
                # Adjust row height to accommodate the image
                ws.row_dimensions[row_num].height = 60
            except Exception as e:
                ws.cell(row=row_num, column=len(df.columns) + 1).value = "Error embedding image"
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        if col == len(headers):  # Logo column
            ws.column_dimensions[chr(64 + col)].width = 25
        else:
            ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save the workbook
    output_filename = "logos_extracted.xlsx"
    wb.save(output_filename)
    return output_filename

def main():
    st.title("Website Logo Scraper")
    st.write("Upload an Excel file with website URLs to extract logos")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
    
    if uploaded_file:
        try:
            # Read the Excel file
            df = pd.read_excel(uploaded_file)
            
            # Check if the DataFrame has any columns
            if df.empty or len(df.columns) == 0:
                st.error("The uploaded file is empty or has no columns.")
                return
            
            # Let the user select the column containing URLs
            url_column = st.selectbox("Select the column containing website URLs", df.columns)
            
            if st.button("Extract Logos"):
                # Add a progress bar
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                # Store the image data
                all_logos = []
                logo_info_dict = {}
                
                # Create new columns for results
                df['logo_found'] = False
                df['logo_filename'] = None
                
                # Process each URL
                for i, row in df.iterrows():
                    url = str(row[url_column])
                    status_text.text(f"Processing {i+1}/{len(df)}: {url}")
                    
                    # Get the logo
                    logo_info = get_site_logo(url)
                    
                    # Update the DataFrame
                    if logo_info:
                        df.at[i, 'logo_found'] = True
                        df.at[i, 'logo_filename'] = logo_info['filename']
                        all_logos.append(logo_info)
                        logo_info_dict[url] = logo_info
                    
                    # Update progress
                    progress_bar.progress((i + 1) / len(df))
                    
                    # Small delay to prevent overwhelming websites
                    time.sleep(0.5)
                
                # Create a zip file with all logos
                if all_logos:
                    zip_filename = "all_logos.zip"
                    with zipfile.ZipFile(zip_filename, 'w') as zipf:
                        for logo in all_logos:
                            zipf.write(logo['path'], logo['filename'])
                    
                    # Provide download button for the zip file
                    with open(zip_filename, "rb") as f:
                        zip_data = f.read()
                        st.download_button(
                            label="Download All Logos (ZIP)",
                            data=zip_data,
                            file_name=zip_filename,
                            mime="application/zip"
                        )
                
                # Create an Excel file with embedded images
                with st.spinner("Creating Excel file with embedded images..."):
                    excel_file = create_excel_with_embedded_images(df, logo_info_dict, url_column)
                
                # Save a simple results Excel without embedded images (as backup)
                simple_excel = "logos_extracted_simple.xlsx"
                df.to_excel(simple_excel, index=False)
                
                # Provide download buttons for the Excel files
                with open(excel_file, "rb") as file:
                    st.download_button(
                        label="Download Excel with Embedded Images",
                        data=file,
                        file_name=excel_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="excel_with_images"
                    )
                    
                with open(simple_excel, "rb") as file:
                    st.download_button(
                        label="Download Simple Excel Results",
                        data=file,
                        file_name=simple_excel,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="excel_simple"
                    )
                
                # Display the results
                success_count = df['logo_found'].sum()
                st.success(f"Processed {len(df)} websites. Successfully extracted {success_count} logos.")
                
                # Show thumbnails of the extracted logos
                if all_logos:
                    st.subheader("Extracted Logos")
                    
                    # Create a grid layout
                    cols = st.columns(3)
                    for i, logo in enumerate(all_logos):
                        col = cols[i % 3]
                        with col:
                            domain = logo['filename'].split('_')[0]
                            st.image(logo['path'], caption=domain, width=150)
                            
                            # Individual download button for each logo
                            with open(logo['path'], "rb") as img_file:
                                img_data = img_file.read()
                                st.download_button(
                                    label="Download",
                                    data=img_data,
                                    file_name=logo['filename'],
                                    mime=f"image/{logo['format']}",
                                    key=f"logo_{i}"
                                )
        
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            st.exception(e)

if __name__ == "__main__":
    main()
